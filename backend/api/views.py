from rest_framework import viewsets, status, permissions, serializers
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from django.db import transaction, IntegrityError, connection
from django.db.models import Q, Prefetch, Count
from django.utils import timezone
from django.http import HttpResponse
import re
from datetime import datetime, timedelta
import json

from .models import User, LawCase, CaseActuacion, CaseAlerta, CaseNote, Cliente, CaseTag, ActuacionTemplate, Aviso
from .serializers import (
    UserSerializer, UserCreateSerializer, UserUpdateSerializer,
    LawCaseSerializer, LawCaseListSerializer, DashboardRecentCaseSerializer,
    CaseActuacionSerializer, CaseAlertaSerializer, DashboardAlertaSerializer,
    CaseNoteSerializer, ClienteSerializer, ClienteMinimalSerializer,
    CaseTagSerializer, ActuacionTemplateSerializer,
    AvisoSerializer, LoginSerializer
)


class AvisoViewSet(viewsets.ModelViewSet):
    queryset = Aviso.objects.filter(active=True)
    serializer_class = AvisoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        # Solo admin puede crear? Por ahora dejemos que logueado, pero en frontend restringimos.
        # Mejor: si no es admin, error?
        # User desire: "solo el admin pueda poner"
        if not self.request.user.is_admin:
             raise permissions.PermissionDenied("Solo administradores pueden publicar avisos.")
        
        # Desactivar anteriores? "Avisos importantes" suena a uno solo o pocos.
        # Vamos a desactivar todos los anteriores para que solo haya UNO vigente si es lo que quiere (tipo banner)
        # O permitir varios. El request dice "el titulo... quiero que sea mas como avisos".
        # Vamos a permitir historial pero el endpoint get por defecto traerá el último.
        # Para el dashboard, usaremos un @action o simplemente listaremos el último en el DashboardView.
        serializer.save(created_by=self.request.user)

    def get_queryset(self):
        return Aviso.objects.filter(active=True).select_related('created_by').order_by('-created_at')


class IsAdminOrReadOnly(permissions.BasePermission):
    """Permiso: solo admin puede acceder"""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        # Para UserViewSet, solo admins pueden acceder (tanto leer como escribir)
        if view.__class__.__name__ == 'UserViewSet':
            return request.user.is_admin
        # Para otros viewsets, solo admin puede escribir
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_admin


class AuthView(APIView):
    """Vista para autenticación"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Login - retorna JWT tokens"""
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'user': UserSerializer(user).data
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CurrentUserView(APIView):
    """Obtener usuario actual"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        return Response(UserSerializer(request.user).data)


class UserListPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 100


class UserViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de usuarios (solo admin)"""
    queryset = User.objects.all()
    permission_classes = [IsAdminOrReadOnly]
    pagination_class = UserListPagination
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        return UserSerializer
    
    def get_queryset(self):
        # Solo los admins pueden ver usuarios
        if not self.request.user.is_authenticated:
            return User.objects.none()
        
        if not self.request.user.is_admin:
            return User.objects.none()
        
        # No permitir eliminar el usuario admin principal
        return User.objects.exclude(id=1)
    
    def perform_create(self, serializer):
        """Crear usuario - solo admins pueden hacerlo"""
        if not self.request.user.is_admin:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo los administradores pueden crear usuarios')
        try:
            serializer.save()
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al crear usuario: {str(e)}, datos: {self.request.data}, usuario actual: {self.request.user.username}")
            # Si es un ValidationError del serializer, re-lanzarlo para que el frontend vea el mensaje
            if hasattr(e, 'detail'):
                raise
            raise serializers.ValidationError({'non_field_errors': [f'Error al crear usuario: {str(e)}']})
    
    def perform_update(self, serializer):
        """Actualizar usuario - solo admins pueden hacerlo"""
        if not self.request.user.is_admin:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo los administradores pueden actualizar usuarios')
        
        # No permitir modificar el usuario admin principal (id=1)
        if serializer.instance and serializer.instance.id == 1:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('No se puede modificar el administrador principal')
        
        try:
            serializer.save()
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al actualizar usuario: {str(e)}, datos: {self.request.data}, usuario actual: {self.request.user.username}")
            if hasattr(e, 'detail'):
                raise
            raise serializers.ValidationError({'non_field_errors': [f'Error al actualizar usuario: {str(e)}']})


class CaseListPagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 50


class LawCaseViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de expedientes"""
    queryset = LawCase.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = CaseListPagination
    
    def get_serializer_class(self):
        if self.action == 'list':
            return LawCaseListSerializer
        return LawCaseSerializer
    
    def get_queryset(self):
        # Optimización Base: Siempre cargar relaciones directas y etiquetas (usadas en listado)
        queryset = LawCase.objects.select_related(
            'created_by', 'last_modified_by', 'cliente'
        ).prefetch_related('etiquetas')

        # Optimización Condicional: Solo cargar datos pesados (actuaciones, alertas, notas) en detalle
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related(
                # Traer actuaciones con sus autores ya cargados
                Prefetch('actuaciones', queryset=CaseActuacion.objects.select_related('created_by', 'last_modified_by').order_by('-fecha', '-created_at')),
                # Traer alertas con sus autores ya cargados
                Prefetch('alertas', queryset=CaseAlerta.objects.select_related('created_by', 'completed_by').order_by('fecha_vencimiento', 'prioridad')),
                # Traer notas con sus autores ya cargados
                Prefetch('notas', queryset=CaseNote.objects.select_related('created_by').order_by('-created_at'))
            )
        
        # Abogados solo ven expedientes donde están asignados como responsable (por username)
        if self.request.user.is_authenticated and getattr(self.request.user, 'rol', None) == 'abogado' and not self.request.user.is_admin:
            queryset = queryset.filter(abogado_responsable=self.request.user.username)
        
        # Filtros avanzados
        search = self.request.query_params.get('search', None)
        estado = self.request.query_params.get('estado', None)
        abogado = self.request.query_params.get('abogado', None)
        fuero = self.request.query_params.get('fuero', None)
        juzgado = self.request.query_params.get('juzgado', None)
        cliente_id = self.request.query_params.get('cliente', None)
        etiqueta_id = self.request.query_params.get('etiqueta', None)
        fecha_inicio_desde = self.request.query_params.get('fecha_inicio_desde', None)
        fecha_inicio_hasta = self.request.query_params.get('fecha_inicio_hasta', None)
        fecha_modificacion_desde = self.request.query_params.get('fecha_modificacion_desde', None)
        fecha_modificacion_hasta = self.request.query_params.get('fecha_modificacion_hasta', None)
        
        if search:
            queryset = queryset.filter(
                Q(caratula__icontains=search) |
                Q(cliente_nombre__icontains=search) |
                Q(nro_expediente__icontains=search) |
                Q(codigo_interno__icontains=search) |
                Q(cliente__nombre_completo__icontains=search) |
                Q(cliente__dni_ruc__icontains=search)
            )
        
        if estado:
            queryset = queryset.filter(estado=estado)
        
        if abogado:
            queryset = queryset.filter(abogado_responsable__icontains=abogado)
        
        if fuero:
            queryset = queryset.filter(fuero=fuero)
        
        if juzgado:
            queryset = queryset.filter(juzgado__icontains=juzgado)
        
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
        
        if etiqueta_id:
            queryset = queryset.filter(etiquetas__id=etiqueta_id)
        
        if fecha_inicio_desde:
            queryset = queryset.filter(fecha_inicio__gte=fecha_inicio_desde)
        
        if fecha_inicio_hasta:
            queryset = queryset.filter(fecha_inicio__lte=fecha_inicio_hasta)
        
        if fecha_modificacion_desde:
            queryset = queryset.filter(updated_at__gte=fecha_modificacion_desde)
        
        if fecha_modificacion_hasta:
            queryset = queryset.filter(updated_at__lte=fecha_modificacion_hasta)
        
        return queryset.distinct() if etiqueta_id else queryset

    def list(self, request, *args, **kwargs):
        """Lista expedientes. Con ?include_clientes=1 devuelve clientes en la misma respuesta (1 round-trip menos)."""
        response = super().list(request, *args, **kwargs)
        if request.query_params.get('include_clientes') == '1':
            clientes_qs = Cliente.objects.only('id', 'nombre_completo').order_by('nombre_completo')[:500]
            response.data['clientes'] = ClienteMinimalSerializer(clientes_qs, many=True).data
        return response

    def perform_create(self, serializer):
        """Generar código interno automáticamente (transacción atómica para evitar race condition)."""
        for attempt in range(3):
            try:
                with transaction.atomic():
                    year = timezone.now().year
                    last = LawCase.objects.order_by('-id').select_for_update().first()
                    if last and last.codigo_interno:
                        match = re.search(r'ENT-(\d+)-', last.codigo_interno)
                        next_num = int(match.group(1)) + 1 if match else LawCase.objects.count() + 1
                    else:
                        next_num = LawCase.objects.count() + 1
                    codigo = f"ENT-{str(next_num).zfill(4)}-{year}-JLCA"
                    
                    serializer.save(
                        codigo_interno=codigo,
                        created_by=self.request.user,
                        last_modified_by=self.request.user
                    )
                return
            except IntegrityError:
                if attempt >= 2:
                    raise
    
    def perform_update(self, serializer):
        serializer.save(last_modified_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def add_actuacion(self, request, pk=None):
        """Agregar actuación a un expediente"""
        caso = self.get_object()
        serializer = CaseActuacionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(caso=caso, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        # Log de errores para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al crear actuación: {serializer.errors}, datos recibidos: {request.data}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def add_alerta(self, request, pk=None):
        """Agregar alerta a un expediente"""
        import logging
        logger = logging.getLogger(__name__)
        caso = self.get_object()
        serializer = CaseAlertaSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(caso=caso, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        logger.warning('add_alerta validation failed: %s', serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def add_note(self, request, pk=None):
        """Agregar nota a un expediente"""
        caso = self.get_object()
        serializer = CaseNoteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(caso=caso, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exportar expedientes a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response(
                {'error': 'openpyxl no está instalado. Ejecuta: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        queryset = self.get_queryset()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expedientes"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            "Código Interno", "Carátula", "Nro. Expediente", "Juzgado", "Fuero",
            "Estado", "Cliente", "DNI/RUC", "Abogado Responsable", "Contraparte",
            "Fecha Inicio", "Última Modificación", "Creado por", "Modificado por"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Datos: iterator() evita cargar todo en memoria; select_related ya aplicado en get_queryset
        row_num = 2
        for caso in queryset.iterator(chunk_size=500):
            cliente_nombre = caso.cliente.nombre_completo if caso.cliente else caso.cliente_nombre
            cliente_dni = caso.cliente.dni_ruc if caso.cliente else caso.cliente_dni
            
            ws.cell(row=row_num, column=1, value=caso.codigo_interno)
            ws.cell(row=row_num, column=2, value=caso.caratula)
            ws.cell(row=row_num, column=3, value=caso.nro_expediente)
            ws.cell(row=row_num, column=4, value=caso.juzgado)
            ws.cell(row=row_num, column=5, value=caso.fuero)
            ws.cell(row=row_num, column=6, value=caso.estado)
            ws.cell(row=row_num, column=7, value=cliente_nombre)
            ws.cell(row=row_num, column=8, value=cliente_dni)
            ws.cell(row=row_num, column=9, value=caso.abogado_responsable)
            ws.cell(row=row_num, column=10, value=caso.contraparte)
            ws.cell(row=row_num, column=11, value=caso.fecha_inicio.strftime('%Y-%m-%d') if caso.fecha_inicio else '')
            ws.cell(row=row_num, column=12, value=caso.updated_at.strftime('%Y-%m-%d %H:%M') if caso.updated_at else '')
            ws.cell(row=row_num, column=13, value=caso.created_by.username if caso.created_by else '')
            ws.cell(row=row_num, column=14, value=caso.last_modified_by.username if caso.last_modified_by else '')
            row_num += 1
        
        # Ajustar ancho de columnas
        column_widths = [18, 40, 18, 25, 12, 12, 30, 15, 25, 30, 12, 18, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Expedientes_Estudio_Neira_Trujillo_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        
        wb.save(response)
        return response

    @action(detail=True, methods=['get'])
    def export_timeline(self, request, pk=None):
        """Exportar timeline del caso (Actuaciones + Alertas) a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response(
                {'error': 'openpyxl no está instalado'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        caso = self.get_object()
        
        # 1. Obtener datos
        actuaciones = caso.actuaciones.all().select_related('created_by')
        alertas = caso.alertas.all().select_related('created_by', 'completed_by')
        
        # 2. Unificar y ordenar cronológicamente
        timeline = []
        for act in actuaciones:
            timeline.append({
                'fecha': act.fecha,
                'hora': None, 
                'tipo_evento': 'ACTUACIÓN',
                'tipo_detalle': act.tipo,
                'descripcion': act.descripcion,
                'estado': 'Realizado',
                'responsable': act.created_by.username if act.created_by else 'Sistema',
                'objeto': act
            })
            
        for al in alertas:
            timeline.append({
                'fecha': al.fecha_vencimiento,
                'hora': al.hora,
                'tipo_evento': 'TAREA / ALERTA',
                'tipo_detalle': al.prioridad,
                'descripcion': f"{al.titulo} - {al.resumen}",
                'estado': 'Cumplido' if al.cumplida else 'Pendiente',
                'responsable': al.created_by.username if al.created_by else 'Sistema',
                'objeto': al
            })
            
        # Ordenar: primero por fecha descendente, luego por hora (si existe)
        timeline.sort(key=lambda x: (x['fecha'] or datetime.min.date(), x['hora'] or datetime.min.time()), reverse=True)
        
        # 3. Generar Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Timeline {caso.codigo_interno}"
        
        # Estilos
        header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid") # Orange branding
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Info del Caso (Header)
        ws.merge_cells('A1:F1')
        ws['A1'] = f"TIMELINE DEL EXPEDIENTE: {caso.codigo_interno} - {caso.caratula}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center
        
        # Encabezados de tabla
        headers = ["Fecha", "Hora", "Tipo Evento", "Detalle / Prioridad", "Descripción / Resumen", "Estado / Responsable"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
            
        # Filas
        row = 4
        for item in timeline:
            ws.cell(row=row, column=1, value=item['fecha']).number_format = 'DD/MM/YYYY'
            ws.cell(row=row, column=2, value=item['hora']).number_format = 'HH:MM'
            ws.cell(row=row, column=3, value=item['tipo_evento'])
            ws.cell(row=row, column=4, value=item['tipo_detalle'])
            ws.cell(row=row, column=5, value=item['descripcion'])
            
            estado_resp = f"{item['estado']} ({item['responsable']})"
            cell_estado = ws.cell(row=row, column=6, value=estado_resp)
            
            # Estilo condicional básico
            if item['tipo_evento'] == 'TAREA / ALERTA' and item['estado'] == 'Pendiente':
                cell_estado.font = Font(color="FF0000", bold=True)
            
            # Bordes para todas las celdas
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
                
            row += 1
            
        # Ajustar anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Timeline_{caso.codigo_interno}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class CaseActuacionViewSet(viewsets.ModelViewSet):
    """ViewSet para actuaciones"""
    queryset = CaseActuacion.objects.select_related('caso', 'created_by', 'last_modified_by')
    serializer_class = CaseActuacionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(last_modified_by=self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset()
        caso_id = self.request.query_params.get('caso', None)
        if caso_id:
            queryset = queryset.filter(caso_id=caso_id)
        return queryset


class CaseAlertaViewSet(viewsets.ModelViewSet):
    """ViewSet para alertas"""
    queryset = CaseAlerta.objects.select_related('caso', 'created_by', 'completed_by')
    serializer_class = CaseAlertaSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = CaseListPagination
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def toggle_cumplida(self, request, pk=None):
        """Toggle estado cumplida de alerta"""
        alerta = self.get_object()
        alerta.cumplida = not alerta.cumplida
        if alerta.cumplida:
            alerta.completed_by = request.user
            alerta.completed_at = timezone.now()
        else:
            alerta.completed_by = None
            alerta.completed_at = None
        alerta.save()
        return Response(self.get_serializer(alerta).data)
    
    def get_queryset(self):
        queryset = super().get_queryset()
        caso_id = self.request.query_params.get('caso', None)
        cumplida = self.request.query_params.get('cumplida', None)
        if caso_id:
            queryset = queryset.filter(caso_id=caso_id)
        if cumplida is not None:
            queryset = queryset.filter(cumplida=cumplida.lower() == 'true')
        return queryset


class CaseNoteViewSet(viewsets.ModelViewSet):
    """ViewSet para notas"""
    queryset = CaseNote.objects.select_related('caso', 'created_by')
    serializer_class = CaseNoteSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset()
        caso_id = self.request.query_params.get('caso', None)
        if caso_id:
            queryset = queryset.filter(caso_id=caso_id)
        return queryset


class ClienteListPagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = 'page_size'
    max_page_size = 500


class ClienteViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de clientes"""
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = ClienteListPagination
    
    def get_queryset(self):
        # Optimización: Annotate total_expedientes en la query principal para evitar N+1 en el serializer
        queryset = Cliente.objects.annotate(total_expedientes_count=Count('expedientes')).order_by('nombre_completo')
        
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(nombre_completo__icontains=search) |
                Q(dni_ruc__icontains=search) |
                Q(email__icontains=search)
            )
        return queryset


class CaseTagViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de etiquetas"""
    queryset = CaseTag.objects.all()
    serializer_class = CaseTagSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        return queryset


class ActuacionTemplateViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de plantillas de actuaciones"""
    queryset = ActuacionTemplate.objects.select_related('created_by')
    serializer_class = ActuacionTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def get_queryset(self):
        queryset = ActuacionTemplate.objects.select_related('created_by')
        tipo = self.request.query_params.get('tipo', None)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        return queryset


class DashboardView(APIView):
    """Vista para datos del dashboard"""
    permission_classes = [permissions.IsAuthenticated]

    @staticmethod
    def _get_cases_queryset_for_user(request):
        """Queryset base de expedientes según rol: admin ve todos, abogado solo los asignados.
        Usa igualdad exacta (no icontains) para aprovechar índice en abogado_responsable."""
        if request.user.is_admin:
            return LawCase.objects.all().order_by('-updated_at')
        return LawCase.objects.filter(
            abogado_responsable=request.user.username
        ).order_by('-updated_at')
    
    def get(self, request):
        """Obtener estadísticas y alertas para dashboard.
        Una sola query raw para stats+fuero+abogado+meses (reduce 4 round-trips a 1)."""
        from django.db.models import Count
        from django.db.models.functions import TruncMonth
        from django.db.models import Q

        def month_shift(dt, delta_months: int):
            year = dt.year + (dt.month - 1 + delta_months) // 12
            month = (dt.month - 1 + delta_months) % 12 + 1
            return dt.replace(year=year, month=month, day=1, hour=0, minute=0, second=0, microsecond=0)

        now = timezone.now()
        month0 = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start_12m = month_shift(month0, -11)
        fuero_order = ['Civil', 'Comercial', 'Penal', 'Laboral', 'Familia']

        # ---- 1 query: aviso ----
        ultimo_aviso = Aviso.objects.filter(active=True).select_related('created_by').order_by('-created_at').first()
        aviso_data = AvisoSerializer(ultimo_aviso).data if ultimo_aviso else None

        # ---- 1 query raw (PostgreSQL): stats+fuero+abogado+meses en 1 round-trip ----
        cases = self._get_cases_queryset_for_user(request)
        if connection.vendor == 'postgresql':
            if request.user.is_admin:
                where_sql, params = "1=1", [start_12m]
            else:
                where_sql, params = "abogado_responsable = %s", [request.user.username, start_12m]
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    WITH base AS (
                        SELECT id, estado, fuero, abogado_responsable, date_trunc('month', created_at)::date as mes
                        FROM api_lawcase WHERE {where_sql}
                    )
                    SELECT
                        (SELECT COUNT(*) FROM base) as total,
                        (SELECT COUNT(*) FROM base WHERE estado = 'Abierto') as open_cases,
                        (SELECT COUNT(*) FROM base WHERE estado = 'En Trámite') as in_progress_cases,
                        (SELECT COUNT(*) FROM base WHERE estado = 'Pausado') as paused_cases,
                        (SELECT COUNT(*) FROM base WHERE estado = 'Cerrado') as closed_cases,
                        (SELECT COALESCE(jsonb_object_agg(fuero, cnt), '{{}}'::jsonb) FROM (
                            SELECT fuero, COUNT(*)::int as cnt FROM base GROUP BY fuero
                        ) f) as fuero_json,
                        (SELECT COALESCE(jsonb_object_agg(abogado_responsable, cnt), '{{}}'::jsonb) FROM (
                            SELECT abogado_responsable, COUNT(*)::int as cnt
                            FROM base WHERE abogado_responsable != '' GROUP BY abogado_responsable
                        ) a) as abogado_json,
                        (SELECT COALESCE(jsonb_object_agg(to_char(mes, 'YYYY-MM'), cnt), '{{}}'::jsonb) FROM (
                            SELECT mes, COUNT(*)::int as cnt FROM base WHERE mes >= %s GROUP BY mes
                        ) m) as month_json
                    FROM (SELECT 1) x
                """, params)
                row = cursor.fetchone()
            total, open_c, in_prog, paused, closed = row[0] or 0, row[1] or 0, row[2] or 0, row[3] or 0, row[4] or 0
            def _to_dict(v):
                if v is None: return {}
                if isinstance(v, dict): return v
                return json.loads(v) if isinstance(v, str) else {}
            fuero_json = _to_dict(row[5])
            abogado_json = _to_dict(row[6])
            month_json = _to_dict(row[7])
        else:
            # SQLite fallback: ORM (4 queries)
            stats_agg = cases.aggregate(
                total=Count('id'),
                open_cases=Count('id', filter=Q(estado=LawCase.CaseStatus.OPEN)),
                in_progress_cases=Count('id', filter=Q(estado=LawCase.CaseStatus.IN_PROGRESS)),
                paused_cases=Count('id', filter=Q(estado=LawCase.CaseStatus.PAUSED)),
                closed_cases=Count('id', filter=Q(estado=LawCase.CaseStatus.CLOSED)),
            )
            total = stats_agg['total'] or 0
            open_c = stats_agg['open_cases'] or 0
            in_prog = stats_agg['in_progress_cases'] or 0
            paused = stats_agg['paused_cases'] or 0
            closed = stats_agg['closed_cases'] or 0
            fuero_json = dict(cases.values('fuero').annotate(cnt=Count('id')).values_list('fuero', 'cnt'))
            abogado_json = dict(cases.exclude(abogado_responsable='').values('abogado_responsable').annotate(cnt=Count('id')).values_list('abogado_responsable', 'cnt'))
            month_rows = cases.filter(created_at__gte=start_12m).annotate(mes=TruncMonth('created_at')).values('mes').annotate(cnt=Count('id'))
            month_json = {r['mes'].strftime('%Y-%m'): r['cnt'] for r in month_rows if r['mes']}

        stats = {'total_cases': total, 'open_cases': open_c, 'in_progress_cases': in_prog, 'paused_cases': paused, 'closed_cases': closed}
        stats_by_fuero = {f: int(fuero_json.get(f, 0)) for f in fuero_order}
        stats_by_abogado = {k: int(v) for k, v in (abogado_json.items() if abogado_json else [])}
        cases_by_month = [
            {'mes': month_shift(month0, -i).strftime('%Y-%m'), 'total': int(month_json.get(month_shift(month0, -i).strftime('%Y-%m'), 0))}
            for i in range(11, -1, -1)
        ]

        # ---- Alertas: subquery sin ORDER BY (más eficiente que IN con miles de IDs) ----
        cases_subquery = cases.only('id').order_by()
        alertas_qs = (
            CaseAlerta.objects.filter(caso__in=cases_subquery)
            .select_related('caso', 'created_by', 'completed_by')
            .order_by('cumplida', 'fecha_vencimiento')
        )[:5]
        alertas_data = DashboardAlertaSerializer(alertas_qs, many=True).data

        # ---- Últimos casos: solo campos usados, sin prefetch etiquetas (1 query menos) ----
        recent_cases = (
            cases.only('id', 'codigo_interno', 'caratula', 'updated_at', 'last_modified_by')
            .select_related('last_modified_by')
        )[:5]

        data = {
            'stats': stats,
            'recent_cases': DashboardRecentCaseSerializer(recent_cases, many=True).data,
            'alertas': alertas_data,
            'cases_by_month': cases_by_month,
            'stats_by_fuero': stats_by_fuero,
            'stats_by_abogado': stats_by_abogado,
            'aviso': aviso_data
        }
        return Response(data)


class DashboardAlertasView(APIView):
    """Alertas paginadas del dashboard, filtradas por expedientes accesibles del usuario."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from rest_framework.pagination import PageNumberPagination

        page = int(request.query_params.get('page', 1))
        page_size = 5

        cases = DashboardView._get_cases_queryset_for_user(request)
        cases_subquery = cases.only('id').order_by()
        alertas_qs = (
            CaseAlerta.objects.filter(caso__in=cases_subquery)
            .select_related('caso', 'created_by', 'completed_by')
            .order_by('cumplida', 'fecha_vencimiento')
        )

        paginator = PageNumberPagination()
        paginator.page_size = page_size
        paginator.page_size_query_param = None
        page_qs = paginator.paginate_queryset(alertas_qs, request)

        return paginator.get_paginated_response(
            DashboardAlertaSerializer(page_qs, many=True).data
        )
